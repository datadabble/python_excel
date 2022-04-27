import pandas as pd
import numpy as np

from openpyxl.workbook import Workbook
from openpyxl import load_workbook 

#create excel workbook and worksheets
wb = Workbook()
ws = wb.active
ws.title = 'First sheet'
ws1 = wb.create_sheet('No this first', 0) #index # determines where new sheet is placed in WB
ws2 = wb.create_sheet('this last')
print(wb.sheetnames)

#loading exisiting WB
wb2 = load_workbook('regions.xlsx')
new_sheet = wb2.create_sheet('new sheet')
active_sheet = wb2.active #must grab active sheet to use it 

cell = active_sheet['A1']
print(cell.value) #if value isnt included prints meta

#print ranges of cells
cell_range = ws['A1': 'C1']
col_c = ws['C'] 
row_range = ws[1 : 5]

#to interate through a matrix
for row in ws.iter_rows(min_row=1, max_col=3, max_row=2, values_only =True): #value only to inculde value not meta
    for cell in row:
        print(cell)