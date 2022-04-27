#parsing large data sets into more workable excels
import pandas as pd
import numpy as np
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook

wb = load_workbook('template.xlsx')
ws = wb.active

#reads dataset and classfies which columns are strings
df = pd.read_csv('crime.csv', encoding='utf-8', dtypes={"INCIDENT_NUMBER": str, "OFFENSE_CODE": str, "OFFENSE_CODE_GROUP": str, "OFFENSE_DESCRIPTION": str, "DISTRICT": str, "REPORTING_AREA": str, "SHOOTING": str, "YEAR": str, "MONTH": str, "DAY_OF_WEEK": str, "HOUR": str})
df1 = df[df["OFFENSE_CODE_GROUP"] == 'Counterfeiting'] #parses df to just counterfeiting offenses. 

df1 = df1.replace(np.nan, 'N/A', regex= True)#removes nan values

total_crimes = len(df.index)
cf_crimes = len(df1.index)
percentage_crimes = (cf_crimes/ total_crimes)* 100
percentage_crimes = round(percentage_crimes, 2)
#addes stats to report template 
ws['O8'].value = total_crimes
ws['P8'].value = cf_crimes
ws['Q8'].value = percentage_crimes
#counts cf crimes by year in each distrct
df1['Count'] = 1
df2 = df1.groupby(['DISTRICT', 'YEAR']).count()['COUNT'].unstack(level=0)#unstack stacks data in cleaner grid format
df2.drop(columns='N/A', inplace=True) #drops NaN columns

rows = dataframe_to_rows(df2) #adds to openpyxl format
#template starts on A8- adjested for that. iterating over template & filling with data
for r_idx, row in enumerate(rows, 8):
    for c_idx, value in enumerate(row, 1):
        ws.cell(row= r_idx, column=c_idx, value=value)

#save to new workbook
ws.save('crime_report.xlsx')

