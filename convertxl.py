import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

wb = load_workbook('regions.xlsx')#loads existing workbook
ws = wb.active
df = pd.read_excel('all_shifts.xlsx')
#create new df of some of the columns from df
df1 = df[('Sales Rep', 'Cost per', 'Units Sold')]
df1['Total'] = df1['Cost per'] * df1['Units Sold']

rows = dataframe_to_rows(df1, index=False) #False so not to paste extra indexes
for r_idx, row in enumerate(rows, 1): #interate through so to print row values not meta, and keeps track of indexs
    for c_idx, col in enumerate(row, 6):
        ws.cell(row=r_idx, column=c_idx, value=col)

wb.save('combind.xlsx')