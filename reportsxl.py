import pandas as pd

df_1 = pd.read_excel('shifts.xlsx', sheet_name='Sheet')
df_2 = pd.read_excel('shifts.xlsx', sheet_name='Sheet')
df_3 = pd.read_excel('shift_3.xlsx')

 #pandas concatanate function works as long as it was the same header columns
df_all = pd.concat([df_1,df_2, df_3], sort=False)

to_excel = df_all.to_excel('all_shifts.xlsx', index=None)

#add a new column and fill out with new values
from openpyxl import load_workbook
from openpyxl.styles import Font

wb= load_workbook('all_shifts.xlsx')
ws = wb.active

total_col = ws['G1']
total_col.font = Font(bold=True)
total_col.value = 'Total'

e_col, f_col = ['E', 'F']
for row in range(2, 300):
    result_cell = 'G{}'.format(row)#stored in pythons string format function
    e_value = ws[e_col + str(row)].value
    f_value = ws[e_col + str(row)].value
    ws[result_cell] = e_value + f_value

wb.save('totaled.xlsx')